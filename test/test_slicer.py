"""
test_slicer.py
==============
Test runner for the probabilistic program slicer.

Lives in test/ — it is an evaluation script, not part of the core
slicing library (src/prob_slicer/).

Usage (run from the repo root):
    python test/test_slicer.py --list
    python test/test_slicer.py --compare
    python test/test_slicer.py --bench olmedo_nontermination --compare
    python test/test_slicer.py --tag key-example --compare
    python test/test_slicer.py --variant ns --save-json results.json
"""

from __future__ import annotations
import argparse, sys, time, json, csv, tracemalloc, resource, platform, io, contextlib
from pathlib import Path

# Repo layout: this file lives in test/, benchmark_loader.py and
# evaluator.py-adjacent utilities live in bench-src/ (a sibling of
# test/), and prob_slicer lives in src/ (installed in editable mode via
# `pip install -e .` — see pyproject.toml). Add bench-src/ to sys.path so
# `from benchmark_loader import ...` below works without requiring
# bench-src/ to be pip-installed too.
REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / 'bench-src'))

# Core slicing library
from prob_slicer import parse, build_cfg, cfg_to_dot, parser, pretty_print
from prob_slicer.ast_nodes import reset_ids
from prob_slicer.dependence import DependenceAnalysis, SliceVariant, set_debug
from prob_slicer.slicer import slice_program, slice_only

# Evaluation utilities: benchmark_loader lives in bench-src/ (see sys.path
# shim above); evaluator lives alongside this file in test/
from benchmark_loader import load_benchmarks, list_benchmarks, BENCHMARKS_DIR
from evaluator import evaluate_benchmark, EvalResult


# ═══════════════════════════════════════════════════════════════════════════════
# Slicing variant metadata
# ═══════════════════════════════════════════════════════════════════════════════

VARIANTS: dict[SliceVariant, dict] = {
    'ns': {
        'label':       'Nontermination-Sensitive, Distribution-Sensitive',
        'cd':          'scd',
        'R':           'obsntd(scd)',
        'description': 'Preserves both termination behaviour and distribution.',
    },
    'nids': {
        'label':       'Nontermination-Insensitive, Distribution-Sensitive',
        'cd':          'wcd',
        'R':           'obsntd(wcd)',
        'description': 'Preserves distribution but not termination probability.',
    },
    'ni': {
        'label':       'Nontermination-Insensitive, Distribution-Insensitive',
        'cd':          'wcd',
        'R':           'obsd(wcd)',
        'description': 'Most aggressive: preserves qualitative output only.',
    },
}


# ═══════════════════════════════════════════════════════════════════════════════
# Memory measurement helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _peak_rss_kb() -> float:
    """
    Return the process-wide peak resident set size (RSS) in KB, i.e. the
    actual OS-level "real" memory the process has used at any point since
    it started (ru_maxrss is a monotonically non-decreasing high-water
    mark for the whole process — it can only go up, never down).

    ru_maxrss units differ by platform:
      - Linux:   kilobytes
      - macOS:   bytes
    This normalizes both to KB so results are comparable across platforms.
    """
    raw = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
    if platform.system() == 'Darwin':
        return raw / 1024.0
    return float(raw)


# ═══════════════════════════════════════════════════════════════════════════════
# Criterion finder
# ═══════════════════════════════════════════════════════════════════════════════

def find_criterion_nodes(cfg) -> set[int]:
    """
    Find the return node(s) in the CFG.
    These form the slicing criterion C per Definition 6.5.

    The return node uses the variables in the return expression.
    The dependency analysis (Definition 6.5) will then find all
    nodes that transitively influence these variables via
    (cd union dd)* and R*.
    """
    from prob_slicer.ast_nodes import CReturn

    return {
        nid for nid, data in cfg.nodes(data=True)
        if isinstance(data.get('ast'), CReturn)
    }

def find_criterion_nodes_extended(cfg, vars: set[str]) -> set[int]:
    """
    Extended criterion: all nodes that USE any variable in vars.
    Not the default — only used when the user explicitly specifies
    additional criterion variables beyond the return expression.
    """
    from prob_slicer.ast_nodes import CReturn

    result = set()
    for nid, data in cfg.nodes(data=True):
        ast = data.get('ast')
        if ast is None:
            continue
        if isinstance(ast, CReturn):
            result.add(nid)
        elif hasattr(ast, 'uses') and ast.uses() & vars:
            result.add(nid)
    return result

def count_while_loops(prog) -> int:
    """Count the number of while loop nodes in a program AST."""
    from prob_slicer.ast_nodes import CWhile, CIf, CSeq, Program

    count = 0
    stack = [prog.body if isinstance(prog, Program) else prog]

    while stack:
        node = stack.pop()
        t    = type(node)

        if t is CWhile:
            count += 1
            stack.append(node.body)

        elif t is CSeq:
            n = node
            while type(n) is CSeq:
                stack.append(n.right)
                n = n.left
            stack.append(n)

        elif t is CIf:
            stack.append(node.then_branch)
            stack.append(node.else_branch)

    return count
# ═══════════════════════════════════════════════════════════════════════════════
# Core runner
# ═══════════════════════════════════════════════════════════════════════════════

def run_benchmark(
    b:           dict,
    variant:     SliceVariant,
    emit_dot:    bool = False,
    phase1_only: bool = False,
    evaluate=False,
    n_eval_runs=10_000
) -> dict:
    """
    Run a single benchmark under a single slicing variant.

    Returns a result dict suitable for printing or JSON export.
    """
    reset_ids()
    tracemalloc.start()
    rss_before_kb = _peak_rss_kb()
    t0 = time.perf_counter()

    # Parse and build CFG
    #print(b['source'])
    prog = parse(b['source'])
    cfg  = build_cfg(prog)


    # Dependence analysis for this variant
    da = DependenceAnalysis(cfg, variant=variant)
    da.compute()

    print(f"  [Slicing]: dependency analyses: Done")

    # Slicing criterion: last def of criterion var + all observe nodes
    criterion   = find_criterion_nodes(cfg)
    slice_nodes = da.slice(criterion)
    
    print(f"  [Slicing]: slice set computation: Done")

    # Reconstruct sliced program
    sliced_prog = (
        slice_only(prog, slice_nodes)
        if phase1_only
        else slice_program(prog, slice_nodes)
    )
    sliced_src = pretty_print(sliced_prog)
    print(f"  [Slicing]: sliced program construction: Done")
    elapsed    = time.perf_counter() - t0

    # Memory usage — two different measurements:
    #
    # 1. tracemalloc: Python-level object allocations traced since
    #    tracemalloc.start() above. current = live bytes at this point,
    #    peak = high-water mark reached anywhere during parsing/CFG-build/
    #    dependence-analysis/slicing. Does NOT capture memory used inside
    #    C extensions outside Python's allocator.
    #
    # 2. resource.getrusage(...).ru_maxrss: process-wide peak resident set
    #    size (RSS) — actual OS-level memory the process has touched,
    #    including C-extension internals. It's a monotonically
    #    non-decreasing high-water mark for the whole process (it can only
    #    go up), so rss_after_kb is the process's peak RSS as of this
    #    point in time, and rss_delta_kb is how much that peak grew during
    #    this phase specifically (0 if the process had already touched at
    #    least this much memory earlier).
    current_mem_bytes, peak_mem_bytes = tracemalloc.get_traced_memory()
    tracemalloc.stop()
    rss_after_kb  = _peak_rss_kb()
    rss_delta_kb  = rss_after_kb - rss_before_kb

    # Metrics
    #all_nodes    = [n for n in cfg.nodes if n not in (0, -1)]
    #program_size = len(all_nodes)
    #slice_size   = len(slice_nodes)

    def count_statements(prog) -> int:
        src = pretty_print(prog)
        return sum(1 for line in src.splitlines() if line.strip())
    program_size = count_statements(prog)
    slice_size   = count_statements(sliced_prog)
    reduction    = 1.0 - slice_size / max(program_size, 1)

    # Emit DOT files if requested
    if emit_dot:
        dot_dir = Path('dot_output')
        dot_dir.mkdir(exist_ok=True)
        (dot_dir / f"{b['name']}_{variant}_cfg.dot").write_text(
            cfg_to_dot(cfg)
        )
        (dot_dir / f"{b['name']}_{variant}_pdg.dot").write_text(
            da.pdg_to_dot()
        )

    result= {
        'name':          b['name'],
        'variant':       variant,
        'description':   b['description'],
        'reference':     b['reference'],
        'criterion_var': b['criterion'],
        'criterion_nid': criterion,
        'tags':          b['tags'],
        'expected':      b['expected'],
        'program_size':  program_size,
        'slice_size':    slice_size,
        'reduction':     reduction,
        'cd_edges':      da.cd.number_of_edges(),
        'dd_edges':      da.ddg.number_of_edges(),
      #  'R_edges':       da.R.number_of_edges(),
        'slice_nodes':   sorted(slice_nodes),
        'sliced_src':    sliced_src,
        'elapsed_ms':    elapsed * 1000,
        # Python-level allocation tracking (tracemalloc)
        'py_peak_memory_kb':    peak_mem_bytes / 1024,
        'py_current_memory_kb': current_mem_bytes / 1024,
        # Process-wide OS-level peak RSS (resource.getrusage), normalized to KB
        'peak_rss_kb':  rss_after_kb,
        'rss_delta_kb': rss_delta_kb,
        'phase1_only':   phase1_only,
        'n_while_loops':count_while_loops(prog)
    }
    if evaluate:
        tracemalloc.start()
        eval_rss_before_kb = _peak_rss_kb()
        eval_result = evaluate_benchmark(
            b, variant, prog, sliced_prog,
            n_runs=n_eval_runs,
        )
        _, eval_peak_bytes = tracemalloc.get_traced_memory()
        tracemalloc.stop()
        eval_rss_after_kb = _peak_rss_kb()
        result['eval'] = eval_result
        result['eval_py_peak_memory_kb'] = eval_peak_bytes / 1024
        result['eval_peak_rss_kb']       = eval_rss_after_kb
        result['eval_rss_delta_kb']      = eval_rss_after_kb - eval_rss_before_kb
        #print(eval_result, flush=True)

    return result



# ═══════════════════════════════════════════════════════════════════════════════
# Reporting
# ═══════════════════════════════════════════════════════════════════════════════

def print_result_old(r: dict):
    vinfo = VARIANTS[r['variant']]
    phase = " [Phase 1 only — before postprocessing]" if r['phase1_only'] else ""
    print(f"\n{'='*65}")
    print(f"  Benchmark : {r['name']}{phase}")
    print(f"  Reference : {r['reference']}")
    print(f"  Tags      : {', '.join(r['tags'])}")
    print(f"  Variant   : {r['variant'].upper()} — {vinfo['label']}")
    print(f"              cd = {vinfo['cd']},  R = {vinfo['R']}")
    print(f"              {vinfo['description']}")
    print(f"  Nodes     : {r['program_size']} total, "
          f"{r['slice_size']} in slice")
    print(f"  Reduction : {r['reduction']*100:.1f}%")
    print(f"  Criterion : node {r['criterion_nid']} "
          f"(var: {r['criterion_var']})")
    #print(f"  Dep edges : cd={r['cd_edges']}, "
    #      f"dd={r['dd_edges']}, R={r['R_edges']}")
    print(f"  Dep edges : cd={r['cd_edges']}, "
          f"dd={r['dd_edges']}")
    print(f"  Time      : {r['elapsed_ms']:.2f} ms")
    print(f"  Expected  : {r['expected']}")
    print(f"{'─'*65}")
    print("  Sliced program:")
    for line in r['sliced_src'].splitlines():
        print(f"    {line}")
    print(f"{'='*65}")


def print_comparison_old(results: list[dict]):
    """Side-by-side comparison of all three variants for one benchmark."""
    if not results:
        return
    name = results[0]['name']
    print(f"\n{'='*65}")
    print(f"  Comparison: {name}")
    print(f"{'─'*65}")
    print(f"  {'Variant':<8} {'Size':>6} {'Reduction':>10} "
          f"{'cd':>5} {'dd':>5} {'R':>5} {'ms':>8}")
    print(f"{'─'*65}")
    for r in results:
        print(f"  {r['variant'].upper():<8} "
              f"{r['slice_size']:>6} "
              f"{r['reduction']*100:>9.1f}% "
              f"{r['cd_edges']:>5} "
              f"{r['dd_edges']:>5} "
              f"{r['R_edges']:>5} "
              f"{r['elapsed_ms']:>8.2f}")
    print(f"{'─'*65}")

    # Check expected ordering: ns >= nids >= ni
    ns   = next((r for r in results if r['variant'] == 'ns'),   None)
    nids = next((r for r in results if r['variant'] == 'nids'), None)
    ni   = next((r for r in results if r['variant'] == 'ni'),   None)
    if ns and nids and ni:
        ok1 = ns['slice_size']   >= nids['slice_size']
        ok2 = nids['slice_size'] >= ni['slice_size']
        print(f"  Ordering ns >= nids : {'✓' if ok1 else '✗'} "
              f"({ns['slice_size']} >= {nids['slice_size']})")
        print(f"  Ordering nids >= ni : {'✓' if ok2 else '✗'} "
              f"({nids['slice_size']} >= {ni['slice_size']})")
    print(f"{'='*65}")


def save_results_json_old(all_results: list[dict], path: str):
    """Save all results to a JSON file for further analysis."""
    clean = [
        {k: v for k, v in r.items() if k != 'sliced_src'}
        for r in all_results
    ]
    with open(path, 'w') as f:
        json.dump(clean, f, indent=2)
    print(f"\n[INFO] Results saved to {path}")


def print_summary_old(all_results: list[dict]):
    """Print a final summary table across all benchmarks and variants."""
    print(f"\n{'='*72}")
    print("  SUMMARY")
    print(f"{'─'*72}")
    print(f"  {'Benchmark':<32} {'Variant':<8} "
          f"{'Size':>6} {'Reduction':>10} {'ms':>8}")
    print(f"{'─'*72}")
    for r in all_results:
        print(f"  {r['name']:<32} {r['variant'].upper():<8} "
              f"{r['slice_size']:>6}/{r['program_size']:>6} "
              f"{r['reduction']*100:>9.1f}% "
              f"{r['elapsed_ms']:>8.2f}")
    print(f"{'='*72}")
    total_ms = sum(r['elapsed_ms'] for r in all_results)
    print(f"  Total experiments : {len(all_results)}")
    print(f"  Total time        : {total_ms:.2f} ms")
    print()

def _eval_str(r: dict) -> str:
    """Format evaluation result inline if present."""
    ev = r.get('eval')
    if ev is None:
        return ''
    def tick(ok): return '✓' if ok else '✗'
    q_str = f"{ev.q:.3f}" if ev.q is not None else "N/A"
    return (
        f"  Correctness:\n"
        f"    Original : term={ev.orig_dist.p_term:.3f}  "
        f"blocked={ev.orig_dist.p_blocked:.3f}  "
        f"diverged={ev.orig_dist.p_diverged:.3f}\n"
        f"    Slice    : term={ev.slice_dist.p_term:.3f}  "
        f"blocked={ev.slice_dist.p_blocked:.3f}  "
        f"diverged={ev.slice_dist.p_diverged:.3f}\n"
        f"    mu1|V = q*mu2|V : q={q_str}  "
        f"TV_shape={ev.tv_shape:.3f}\n"
        f"    NT linear fit   : q1={ev.q1:.3f}  q2={ev.q2:.3f}  "
        f"NT_diff={ev.nt_diff:.3f}\n"
        f"    NS={tick(ev.ns_ok)}  "
        f"NI={tick(ev.ni_ok)}  "
        f"NIDS={tick(ev.nids_ok)}  "
        f"({ev.n_runs} runs, {ev.elapsed:.2f}s)"
    )


def print_result(r: dict):
    vinfo = VARIANTS[r['variant']]
    phase = " [Phase 1 only — before postprocessing]" if r['phase1_only'] else ""
    print(f"\n{'='*65}")
    print(f"  Benchmark : {r['name']}{phase}")
    print(f"  Reference : {r['reference']}")
    print(f"  Tags      : {', '.join(r['tags'])}")
    print(f"  Variant   : {r['variant'].upper()} — {vinfo['label']}")
    print(f"              cd = {vinfo['cd']},  R = {vinfo['R']}")
    print(f"              {vinfo['description']}")
    print(f"  Nodes     : {r['program_size']} total, "
          f"{r['slice_size']} in slice  "
          f"(while loops: {r.get('n_while_loops', 0)})")
    print(f"  Reduction : {r['reduction']*100:.1f}%")
    print(f"  Criterion : node {r['criterion_nid']} "
          f"(var: {r['criterion_var']})")
    print(f"  Dep edges : cd={r['cd_edges']}, "
          f"dd={r['dd_edges']}")
    print(f"  Time      : {r['elapsed_ms']:.2f} ms")
    print(f"  Peak RSS  : {r.get('peak_rss_kb', 0):.1f} KB "
          f"(process, slicing phase; +{r.get('rss_delta_kb', 0):.1f} KB growth)"
          + (f", {r['eval_peak_rss_kb']:.1f} KB (process, evaluation phase; "
             f"+{r.get('eval_rss_delta_kb', 0):.1f} KB growth)"
             if r.get('eval_peak_rss_kb') is not None else ""))
    print(f"  Py alloc  : {r.get('py_peak_memory_kb', 0):.1f} KB peak (slicing)"
          + (f", {r['eval_py_peak_memory_kb']:.1f} KB peak (evaluation)"
             if r.get('eval_py_peak_memory_kb') is not None else ""))
    print(f"  Expected  : {r['expected']}")
    ev_str = _eval_str(r)
    if ev_str:
        print(f"{'─'*65}")
        print(ev_str)
    print(f"{'─'*65}")
    print("  Sliced program:")
    for line in r['sliced_src'].splitlines():
        print(f"    {line}")
    print(f"{'='*65}")


def print_comparison(results: list[dict]):
    """Side-by-side comparison of all three variants for one benchmark."""
    if not results:
        return
    name    = results[0]['name']
    has_eval = any(r.get('eval') is not None for r in results)

    print(f"\n{'='*65}")
    print(f"  Comparison: {name}")
    print(f"{'─'*65}")

    if has_eval:
        print(f"  {'Variant':<8} {'Size':>6} {'Reduction':>10} "
              f"{'cd':>5} {'dd':>5} {'ms':>8} {'RSS_KB':>9} {'PY_KB':>9} "
              f"{'NS':>4} {'NI':>4} {'NIDS':>6} {'q':>6} {'TV':>6}")
    else:
        print(f"  {'Variant':<8} {'Size':>6} {'Reduction':>10} "
              f"{'cd':>5} {'dd':>5} {'ms':>8} {'RSS_KB':>9} {'PY_KB':>9}")
    print(f"{'─'*65}")

    for r in results:
        ev = r.get('eval')
        def tick(ok): return '✓' if ok else '✗'
        base = (
            f"  {r['variant'].upper():<8} "
            f"{r['slice_size']:>6} "
            f"{r['reduction']*100:>9.1f}% "
            f"{r.get('n_while_loops', 0):>6} "
            f"{r['cd_edges']:>5} "
            f"{r['dd_edges']:>5} "
            f"{r['elapsed_ms']:>8.2f} "
            f"{r.get('peak_rss_kb', 0):>9.1f} "
            f"{r.get('py_peak_memory_kb', 0):>9.1f}"
        )
        if ev is not None:
            q_str = f"{ev.q:.3f}" if ev.q is not None else " N/A"
            base += (
                f" {tick(ev.ns_ok):>4}"
                f" {tick(ev.ni_ok):>4}"
                f" {tick(ev.nids_ok):>6}"
                f" {q_str:>6}"
                f" {ev.tv_shape:>6.3f}"
            )
        print(base)

    print(f"{'─'*65}")

    # Ordering check
    ns   = next((r for r in results if r['variant'] == 'ns'),   None)
    nids = next((r for r in results if r['variant'] == 'nids'), None)
    ni   = next((r for r in results if r['variant'] == 'ni'),   None)
    if ns and nids and ni:
        ok1 = ns['slice_size']   >= nids['slice_size']
        ok2 = nids['slice_size'] >= ni['slice_size']
        print(f"  Ordering ns >= nids : {'✓' if ok1 else '✗'} "
              f"({ns['slice_size']} >= {nids['slice_size']})")
        print(f"  Ordering nids >= ni : {'✓' if ok2 else '✗'} "
              f"({nids['slice_size']} >= {ni['slice_size']})")
    print(f"{'='*65}")


def print_summary(all_results: list[dict]):
    """Print a final summary table across all benchmarks and variants."""
    has_eval = any(r.get('eval') is not None for r in all_results)

    print(f"\n{'='*72}")
    print("  SUMMARY")
    print(f"{'─'*72}")

    if has_eval:
        print(f"  {'Benchmark':<32} {'Variant':<8} "
              f"{'Size':>6} {'While':>6} {'Reduction':>10} {'ms':>8} "
              f"{'RSS_KB':>9} {'PY_KB':>9} "
              f"{'NS':>4} {'NI':>4} {'NIDS':>6}")
    else:
        print(f"  {'Benchmark':<32} {'Variant':<8} "
              f"{'Size':>6} {'While':>6} {'Reduction':>10} {'ms':>8} "
              f"{'RSS_KB':>9} {'PY_KB':>9}")
    print(f"{'─'*72}")

    for r in all_results:
        ev = r.get('eval')
        def tick(ok): return '✓' if ok else '✗'
        base = (
            f"  {r['name']:<32} {r['variant'].upper():<8} "
            f"{r['slice_size']:>6}/{r['program_size']:>6} "
            f"{r.get('n_while_loops', 0):>6} "
            f"{r['reduction']*100:>9.1f}% "
            f"{r['elapsed_ms']:>8.2f} "
            f"{r.get('peak_rss_kb', 0):>9.1f} "
            f"{r.get('py_peak_memory_kb', 0):>9.1f}"
        )
        if ev is not None:
            base += (
                f" {tick(ev.ns_ok):>4}"
                f" {tick(ev.ni_ok):>4}"
                f" {tick(ev.nids_ok):>6}"
            )
        print(base)

    print(f"{'─'*72}")
    total_ms   = sum(r['elapsed_ms'] for r in all_results)
    n_while   = sum(r.get('n_while_loops', 0) for r in all_results)
    max_rss_kb = max((r.get('peak_rss_kb', 0) for r in all_results), default=0)
    avg_rss_kb = (
        sum(r.get('peak_rss_kb', 0) for r in all_results) / len(all_results)
        if all_results else 0
    )
    max_py_kb = max((r.get('py_peak_memory_kb', 0) for r in all_results), default=0)
    avg_py_kb = (
        sum(r.get('py_peak_memory_kb', 0) for r in all_results) / len(all_results)
        if all_results else 0
    )

    total_exp  = len(all_results)

    if has_eval:
        ns_ok   = sum(1 for r in all_results
                      if r.get('eval') and r['eval'].ns_ok)
        ni_ok   = sum(1 for r in all_results
                      if r.get('eval') and r['eval'].ni_ok)
        nids_ok = sum(1 for r in all_results
                      if r.get('eval') and r['eval'].nids_ok)
        n_eval  = sum(1 for r in all_results if r.get('eval'))
        print(f"  Correctness (out of {n_eval} evaluated):")
        print(f"    NS   passed: {ns_ok}/{n_eval}")
        print(f"    NI   passed: {ni_ok}/{n_eval}")
        print(f"    NIDS passed: {nids_ok}/{n_eval}")

    # Memory usage broken down per variant (peak + average across all
    # benchmarks run under that variant, for both memory metrics)
    print(f"{'─'*72}")
    print("  Memory usage by variant (peak / avg across benchmarks):")
    variants_seen = sorted({r['variant'] for r in all_results})
    for v in variants_seen:
        rows = [r for r in all_results if r['variant'] == v]
        v_max_rss = max((r.get('peak_rss_kb', 0) for r in rows), default=0)
        v_avg_rss = sum(r.get('peak_rss_kb', 0) for r in rows) / len(rows)
        v_max_py  = max((r.get('py_peak_memory_kb', 0) for r in rows), default=0)
        v_avg_py  = sum(r.get('py_peak_memory_kb', 0) for r in rows) / len(rows)
        print(f"    {v.upper():<6}  "
              f"RSS: peak={v_max_rss:>9.1f} KB  avg={v_avg_rss:>9.1f} KB  |  "
              f"Py alloc: peak={v_max_py:>9.1f} KB  avg={v_avg_py:>9.1f} KB  "
              f"(n={len(rows)})")

    print(f"{'─'*72}")
    print(f"  Total while loops : {n_while}")

    print(f"  Total experiments : {total_exp}")
    print(f"  Total time        : {total_ms:.2f} ms")
    print(f"  Peak RSS          : {max_rss_kb:.1f} KB (max), "
          f"{avg_rss_kb:.1f} KB (avg per experiment)")
    print(f"  Peak Py alloc     : {max_py_kb:.1f} KB (max), "
          f"{avg_py_kb:.1f} KB (avg per experiment)")
    print()


def summary_to_text(all_results: list[dict]) -> str:
    """Render print_summary()'s output as a string instead of printing it."""
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        print_summary(all_results)
    return buf.getvalue()


def save_summary_txt(all_results: list[dict], path: str, header: str | None = None):
    """
    Save the --summary table (and memory/correctness breakdown) to a
    plain-text file, so it's not console-output-only.

    If `header` is given, it's written as a line above the summary (used
    by run_all_benchmarks.py to label which benchmark directory a section
    belongs to when appending multiple summaries to one file).
    """
    text = summary_to_text(all_results)
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, 'w') as f:
        if header:
            f.write(f"{header}\n")
        f.write(text)
    print(f"\n[INFO] Summary saved to {out_path}")


def save_results_json(all_results: list[dict], path: str):
    """Save all results to a JSON file for further analysis."""
    def serialise(r: dict) -> dict:
        clean = {k: v for k, v in r.items() if k != 'sliced_src'}
        if 'criterion_nid' in clean and isinstance(clean['criterion_nid'], set):
            clean['criterion_nid'] = sorted(clean['criterion_nid'])
        ev = clean.pop('eval', None)
        if ev is not None:
            clean['eval'] = {
                'ns_ok':        ev.ns_ok,
                'ni_ok':        ev.ni_ok,
                'nids_ok':      ev.nids_ok,
                'q':            ev.q,
                'tv_shape':     ev.tv_shape,
                'nt_diff':      ev.nt_diff,
                'q1':           ev.q1,
                'q2':           ev.q2,
                'orig_term':    ev.orig_dist.p_term,
                'orig_blocked': ev.orig_dist.p_blocked,
                'orig_nt':      ev.orig_dist.p_diverged,
                'slice_term':   ev.slice_dist.p_term,
                'slice_blocked':ev.slice_dist.p_blocked,
                'slice_nt':     ev.slice_dist.p_diverged,
                'n_runs':       ev.n_runs,
                'elapsed':      ev.elapsed,
            }
        return clean

    with open(path, 'w') as f:
        json.dump([serialise(r) for r in all_results], f, indent=2)
    print(f"\n[INFO] Results saved to {path}")


def save_results_csv(all_results: list[dict], path: str):
    """
    Save all results to a flat CSV file for spreadsheet analysis.

    Non-scalar fields (sliced_src, slice_nodes, tags, criterion_nid) are
    either dropped or joined into a single string column. The `eval`
    sub-object (present only when --evaluate was used) is flattened into
    eval_* columns.
    """
    def flatten(r: dict) -> dict:
        row = {
            'name':          r.get('name'),
            'variant':       r.get('variant'),
            'description':   r.get('description'),
            'reference':     r.get('reference'),
            'criterion_var': r.get('criterion_var'),
            'criterion_nid': ';'.join(str(n) for n in sorted(r.get('criterion_nid', []))),
            'tags':          ';'.join(r.get('tags', [])),
            'expected':      r.get('expected'),
            'program_size':  r.get('program_size'),
            'slice_size':    r.get('slice_size'),
            'reduction':     r.get('reduction'),
            'cd_edges':      r.get('cd_edges'),
            'dd_edges':      r.get('dd_edges'),
            'elapsed_ms':    r.get('elapsed_ms'),
            'peak_rss_kb':          r.get('peak_rss_kb'),
            'rss_delta_kb':         r.get('rss_delta_kb'),
            'py_peak_memory_kb':    r.get('py_peak_memory_kb'),
            'py_current_memory_kb': r.get('py_current_memory_kb'),
            'phase1_only':   r.get('phase1_only'),
            'n_while_loops': r.get('n_while_loops'),
        }
        ev = r.get('eval')
        if ev is not None:
            row.update({
                'eval_ns_ok':        ev.ns_ok,
                'eval_ni_ok':        ev.ni_ok,
                'eval_nids_ok':      ev.nids_ok,
                'eval_q':            ev.q,
                'eval_tv_shape':     ev.tv_shape,
                'eval_nt_diff':      ev.nt_diff,
                'eval_q1':           ev.q1,
                'eval_q2':           ev.q2,
                'eval_orig_term':    ev.orig_dist.p_term,
                'eval_orig_blocked': ev.orig_dist.p_blocked,
                'eval_orig_nt':      ev.orig_dist.p_diverged,
                'eval_slice_term':   ev.slice_dist.p_term,
                'eval_slice_blocked':ev.slice_dist.p_blocked,
                'eval_slice_nt':     ev.slice_dist.p_diverged,
                'eval_n_runs':       ev.n_runs,
                'eval_elapsed':      ev.elapsed,
                'eval_peak_rss_kb':       r.get('eval_peak_rss_kb'),
                'eval_rss_delta_kb':      r.get('eval_rss_delta_kb'),
                'eval_py_peak_memory_kb': r.get('eval_py_peak_memory_kb'),
            })
        return row

    rows = [flatten(r) for r in all_results]
    if not rows:
        print("\n[WARNING] No results to save to CSV.")
        return

    # Union of all keys across rows, preserving first-seen order
    # (rows without --evaluate won't have eval_* columns)
    fieldnames: list[str] = []
    for row in rows:
        for k in row:
            if k not in fieldnames:
                fieldnames.append(k)

    with open(path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"\n[INFO] Results saved to {path}")


def save_results_xlsx(all_results: list[dict], path: str, sheet_name: str):
    """
    Write/overwrite one sheet of an Excel workbook with a flat results
    table, one row per benchmark x variant.

    This is the automated replacement for the old workflow of manually
    pasting console output into results/experimental-result.xlsx: run
    test_slicer.py (or run_all_benchmarks.py) with --save-xlsx and the
    resulting workbook can be fed straight into data_analysis.py /
    get_statistics.py, which auto-discover sheets and treat each sheet
    name as a benchmark category (typically one sheet per benchmarks/
    subdirectory, e.g. 'prodigy', 'real-world', 'literature', ...).

    If `path` already exists, only `sheet_name` is replaced — other
    sheets in the workbook are left untouched. If `path` doesn't exist,
    a new workbook is created.

    Columns: Benchmark, Variant, orig, slice, While, Reduction, ms,
    RSS_KB (process-wide peak RSS, the actual memory-usage metric),
    PY_KB (tracemalloc peak Python-level allocation, supplementary),
    NS, NI, NIDS (NS/NI/NIDS are '✓'/'✗' correctness flags from
    --evaluate; left blank if --evaluate wasn't used).
    """
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        print("\n[ERROR] --save-xlsx requires the 'openpyxl' package. "
              "Install it with: pip install openpyxl")
        return

    xlsx_path = Path(path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        default = wb.active
        if default.title == 'Sheet':
            wb.remove(default)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    header = ['Benchmark', 'Variant', 'orig', 'slice', 'While',
              'Reduction', 'ms', 'RSS_KB', 'PY_KB', 'NS', 'NI', 'NIDS']
    ws.append(header)

    def tick(ok) -> str:
        if ok is None:
            return ''
        return '✓' if ok else '✗'

    for r in all_results:
        ev = r.get('eval')
        ws.append([
            r.get('name'),
            str(r.get('variant', '')).upper(),
            r.get('program_size'),
            r.get('slice_size'),
            r.get('n_while_loops', 0),
            r.get('reduction'),
            r.get('elapsed_ms'),
            r.get('peak_rss_kb'),
            r.get('py_peak_memory_kb'),
            tick(ev.ns_ok)   if ev is not None else '',
            tick(ev.ni_ok)   if ev is not None else '',
            tick(ev.nids_ok) if ev is not None else '',
        ])

    wb.save(xlsx_path)
    print(f"\n[INFO] Results saved to sheet '{sheet_name}' in {xlsx_path}")

# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    ap = argparse.ArgumentParser(
        description=(
            'Test the probabilistic program slicer on benchmarks '
            'from the benchmarks/ directory.'
        )
    )
    ap.add_argument(
        '--bench', default=None, nargs='+',
        metavar='NAME',
        help='Run specific benchmarks by name (default: all)'
    )
    ap.add_argument(
        '--tag', default=None, nargs='+',
        metavar='TAG',
        help='Run only benchmarks matching these tags'
    )
    ap.add_argument(
        '--benchdir', default=None, metavar='DIR',
        help='Directory to load .prob benchmark files from '
             '(default: benchmarks/real-world/ next to benchmark_loader.py)'
    )
    ap.add_argument(
        '--variant', default=None,
        choices=['ns', 'nids', 'ni'],
        help='Run a single slicing variant (default: all three)'
    )
    ap.add_argument(
        '--compare', action='store_true',
        help='Print side-by-side comparison of all three variants '
             'per benchmark'
    )
    ap.add_argument(
        '--summary', action='store_true',
        help='Print a summary table of all results at the end'
    )
    ap.add_argument(
        '--phase1', action='store_true',
        help='Show Phase 1 output (Definition 6.6, before postprocessing)'
    )
    ap.add_argument(
        '--dot', action='store_true',
        help='Emit .dot files for CFG and dependence graphs '
             'into dot_output/ directory'
    )
    ap.add_argument(
        '--verbose', action='store_true',
        help='Print full dependence report for each experiment'
    )
    ap.add_argument(
        '--list', action='store_true',
        help='List all available benchmarks and exit'
    )
    ap.add_argument(
        '--save-json', default=None, metavar='FILE',
        help='Save all results (excluding sliced source) to a JSON file'
    )
    ap.add_argument(
        '--save-csv', default=None, metavar='FILE',
        help='Save all results (flattened, excluding sliced source) '
             'to a CSV file'
    )
    ap.add_argument(
        '--save-xlsx', default=None, metavar='FILE',
        help='Write/overwrite a sheet of results in an Excel workbook '
             '(pairs with --sheet-name). Only the target sheet is '
             'touched; other sheets in an existing workbook are kept. '
             'Requires openpyxl.'
    )
    ap.add_argument(
        '--sheet-name', default=None, metavar='NAME',
        help='Sheet name to use with --save-xlsx (default: the '
             '--benchdir directory name, or "results" if --benchdir '
             'was not given)'
    )
    ap.add_argument(
        '--save-summary-txt', default=None, metavar='FILE',
        help='Save the --summary table (with per-variant memory '
             'breakdown) to a plain-text file. Requires --summary.'
    )
    ap.add_argument('--debug', action='store_true',
                    help='Enable slice computation progress output')
    ap.add_argument('--evaluate', action='store_true',
                    help='Run Monte Carlo correctness evaluation')
    ap.add_argument('--eval-runs', type=int, default=10_000,
                    help='Number of Monte Carlo runs for evaluation')
    args = ap.parse_args()
    set_debug(args.debug)

    # List benchmarks and exit
    if args.list:
        list_benchmarks(benchmarks_dir=args.benchdir)
        sys.exit(0)

    # Load benchmarks from --benchdir if given, else the default
    # benchmarks/real-world/ directory
    try:
        benchmarks = load_benchmarks(
            tags=args.tag,
            names=args.bench,
            benchmarks_dir=args.benchdir,
        )
    except FileNotFoundError as e:
        print(f"[ERROR] {e}")
        sys.exit(1)

    if not benchmarks:
        print("[ERROR] No benchmarks matched the given filters.")
        print("        Use --list to see all available benchmarks.")
        sys.exit(1)

    bench_dir_used = Path(args.benchdir) if args.benchdir else BENCHMARKS_DIR
    print(f"\n[INFO] Loaded {len(benchmarks)} benchmark(s) "
          f"from {bench_dir_used.resolve()}")

    variants: list[SliceVariant] = (
        [args.variant] if args.variant
        else ['ns', 'nids', 'ni']
    )

    all_results: list[dict] = []

    for b in benchmarks:
        print(f"\n[INFO] Running benchmark: {b['name']} ")
        per_benchmark_results = []

        for variant in variants:
            print(f"\n[INFO] Variant: {variant.upper()}")
            try:
                r = run_benchmark(
                    b,
                    variant=variant,
                    emit_dot=args.dot,
                    phase1_only=args.phase1,
                    evaluate=args.evaluate,
                    n_eval_runs=args.eval_runs,
                )
                per_benchmark_results.append(r)
                all_results.append(r)
                print_result(r)

                if args.verbose:
                    reset_ids()
                    prog = parse(b['source'])
                    cfg  = build_cfg(prog)
                    da   = DependenceAnalysis(cfg, variant=variant)
                    da.compute()
                    da.print_report()

            except Exception as e:
                print(f"\n[ERROR] {b['name']!r} variant={variant}: {e}")
                import traceback; 
                error_string = traceback.format_exc()
                print(error_string)
                #traceback.print_exc()

        if args.compare and len(per_benchmark_results) > 1:
            print_comparison(per_benchmark_results)

    if args.summary:
        print_summary(all_results)
        if args.save_summary_txt:
            save_summary_txt(all_results, args.save_summary_txt)
    elif args.save_summary_txt:
        print("[WARNING] --save-summary-txt requires --summary; nothing saved.")

    if args.save_json:
        save_results_json(all_results, args.save_json)

    if args.save_csv:
        save_results_csv(all_results, args.save_csv)

    if args.save_xlsx:
        sheet_name = (
            args.sheet_name
            or (Path(args.benchdir).name if args.benchdir else 'results')
        )
        save_results_xlsx(all_results, args.save_xlsx, sheet_name)

    print(f"\n[DONE] {len(all_results)} experiment(s) completed "
          f"({len(benchmarks)} benchmark(s) x "
          f"{len(variants)} variant(s)).")